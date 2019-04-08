from openpyxl import load_workbook

#打开一个workbook
wb1 = load_workbook(filename="1.xlsx")
wb2 = load_workbook(filename="2.xlsx")

ws1 = wb1.active
ws2 = wb2.active

for i in range(1, 4500):
    A = ws1.cell(row=i, column=1).value
    if A == None:
        break
    for j in range(1, 4500):
        B = ws2.cell(row=j, column=1).value
        if A == B:
            ws1.cell(row=i, column=2).value = ws2.cell(row=j, column=2).value
        if B == None:
            break
            
print(str(i) + ' rows processed in 1.xlsx.')
print(str(j) + ' rows processed in 2.xlsx.')

if i >=4500 or j >= 4500:
    print('The input data exceeds 4500 rows. The process is terminated to prevent 當機.')
            
wb1.save(filename="1.xlsx")